import datetime
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from social.models import Client, Salarie


def clean_nir(value):
    """
    Nettoie le NIR (Numéro de sécu) :
    - gère les valeurs numériques (float/int)
    - supprime les espaces
    """
    if value is None:
        return ""

    # Si Excel l'a stocké sous forme de nombre (float / int)
    if isinstance(value, (int, float)):
        return str(int(value))

    s = str(value).strip().replace(" ", "")
    return s


class Command(BaseCommand):
    help = "Import des salariés depuis un fichier Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "xlsx_path",
            help="Chemin du fichier Excel des salariés (ex: LISTE SALARIES 2AH.xlsx)",
        )
        parser.add_argument(
            "--raison",
            help="Raison sociale / nomDossier du client (sinon récupérée en E1 du fichier)",
            default=None,
        )

    def handle(self, *args, **options):
        xlsx_path = options["xlsx_path"]
        raison = options["raison"]

        self.stdout.write(f"Import des salariés depuis : {xlsx_path}")

        # 1️⃣ Ouverture du fichier Excel
        try:
            wb = load_workbook(xlsx_path, data_only=True)
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f"Fichier introuvable : {xlsx_path}"))
            return

        ws = wb.active

        # 2️⃣ Récupération de la raison sociale / nom dossier
        if not raison:
            raison = ws["E1"].value  # ligne 1, colonne E = '2 AH' dans ton fichier

        if not raison:
            self.stdout.write(self.style.ERROR("Impossible de déterminer la raison sociale (cellule E1 vide)."))
            return

        self.stdout.write(f"Entreprise ciblée : {raison}")

        # On essaie d'abord sur raison_social, puis sur nom_dossier
        try:
            client = Client.objects.get(raison_social=raison)
        except Client.DoesNotExist:
            try:
                client = Client.objects.get(nom_dossier=raison)
            except Client.DoesNotExist:
                self.stdout.write(
                    self.style.ERROR(
                        f"Aucun client trouvé avec raison_social ou nom_dossier = '{raison}'"
                    )
                )
                return

        # 3️⃣ Ligne d'en-tête et début des données
        # Dans TON fichier :
        # - ligne 7 = en-têtes: NOM, PRENOM, ...
        # - ligne 10 = première ligne de salarié
        header_row_index = 7   # ✅ ligne avec NOM, PRENOM, etc.
        data_start_row = 10    # ✅ première ligne de données (ACHIK SAID)



        headers = [cell.value for cell in ws[header_row_index]]
        # On construit un dict { "NOM": index_col, ... }
        header_index = {name: idx for idx, name in enumerate(headers) if name}

        def get(row, col_name):
            idx = header_index.get(col_name)
            if idx is None:
                return None
            # openpyxl est 1-based, enumerate 0-based
            return ws.cell(row=row, column=idx + 1).value

        created = 0
        updated = 0

        # 4️⃣ Boucle sur chaque salarié
        for row in range(data_start_row, ws.max_row + 1):
            nom = get(row, "NOM")
            prenom = get(row, "PRENOM")

            # Si la ligne est vide, on arrête (on suppose fin de fichier)
            if not nom and not prenom:
                continue

            numero_ss = clean_nir(get(row, "NumeroSS"))

            date_naissance = get(row, "DateNaissance")
            if isinstance(date_naissance, datetime.datetime):
                date_naissance = date_naissance.date()

            date_entree = get(row, "DateEntree")
            if isinstance(date_entree, datetime.datetime):
                date_entree = date_entree.date()

            salarie, is_created = Salarie.objects.update_or_create(
                client=client,
                numero_securite_sociale=numero_ss,
                defaults={
                    "nom": nom,
                    "prenom": prenom,
                    "nom_marital": get(row, "NomMarital") or "",
                    "code_civilite": get(row, "CodeCivilite") or "",

                    "date_naissance": date_naissance,
                    "commune_naissance": get(row, "NomCommune") or "",
                    "code_commune_naissance": str(get(row, "CodeCommune") or "") or "",
                    "pays_naissance": get(row, "PaysNaissance") or "",

                    "nationalite": get(row, "Nationalite") or "",

                    "num_voie": str(get(row, "NumVoieSal") or "") or "",
                    "nom_voie": get(row, "NatNomVoieSal") or "",
                    "code_postal": str(get(row, "CodePostalSalarie") or "") or "",
                    "ville": get(row, "burdistSalVille") or "",

                    "code_reglement": get(row, "CodeReglement") or "",
                    "emploi": get(row, "Emploi") or "",
                    "qualification": get(row, "Qualification") or "",
                    "date_entree": date_entree,
                },
            )

            if is_created:
                created += 1
                self.stdout.write(f"Créé : {nom} {prenom} ({numero_ss})")
            else:
                updated += 1
                self.stdout.write(f"Mise à jour : {nom} {prenom} ({numero_ss})")

        self.stdout.write(
            self.style.SUCCESS(
                f"Import terminé. {created} créés, {updated} mis à jour pour l'entreprise {raison}."
            )
        )
